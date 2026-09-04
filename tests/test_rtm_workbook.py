import unittest
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
WORKBOOK = ROOT / "第14小组_基于大模型的个人知识库系统_需求跟踪矩阵.xlsx"


class RequirementTraceabilityMatrixTest(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.workbook = load_workbook(WORKBOOK, data_only=False)
        cls.rtm = cls.workbook["设计用RTM"]
        cls.history = cls.workbook["变更履历"]

    def test_workbook_has_wenkb_sheets_and_no_template_project_content(self):
        self.assertIn("需求跟踪矩阵封面", self.workbook.sheetnames)
        self.assertIn("变更履历", self.workbook.sheetnames)
        self.assertIn("设计用RTM", self.workbook.sheetnames)

        all_values = [
            str(cell.value)
            for worksheet in self.workbook.worksheets
            for row in worksheet.iter_rows()
            for cell in row
            if cell.value is not None
        ]
        content = "\n".join(all_values)
        self.assertNotIn("旧衣回收", content)
        self.assertNotIn("积分商城", content)
        self.assertIn("WenKB", content)

    def test_rtm_contains_all_baseline_requirements_and_statuses(self):
        expected_ids = [
            *(f"FR-{index:03d}" for index in range(1, 19)),
            *(f"NFR-{index:03d}" for index in range(1, 8)),
        ]
        actual_ids = [self.rtm.cell(row, 2).value for row in range(6, 31)]
        self.assertEqual(expected_ids, actual_ids)

        allowed_statuses = {"○", "△", "×", "N/A"}
        for row in range(6, 31):
            self.assertTrue(self.rtm.cell(row, 5).value)
            self.assertTrue(self.rtm.cell(row, 6).value)
            for column in (7, 8, 9):
                self.assertIn(self.rtm.cell(row, column).value, allowed_statuses)
            self.assertTrue(self.rtm.cell(row, 10).value)
            self.assertTrue(self.rtm.cell(row, 11).value)

    def test_change_history_has_twenty_numbered_records_and_key_changes(self):
        records = []
        for row in range(3, 23):
            sequence = self.history.cell(row, 2).value
            self.assertEqual(row - 2, sequence)
            change_date = self.history.cell(row, 3).value
            self.assertIsInstance(change_date, (date, datetime))
            normalized_date = change_date.date() if isinstance(change_date, datetime) else change_date
            self.assertLessEqual(normalized_date, date(2026, 9, 4))
            self.assertTrue(self.history.cell(row, 4).value)
            self.assertTrue(self.history.cell(row, 5).value)
            self.assertEqual(self.history.cell(row, 12).value, 25)
            records.append(str(self.history.cell(row, 4).value))

        content = "\n".join(records)
        for keyword in ("健康检查", "引用", "AES", "OpenAI", "embedding", "架构与知识点"):
            self.assertIn(keyword, content)

    def test_history_summary_and_workbook_calculation_mode_are_present(self):
        self.assertEqual(self.history.cell(23, 2).value, "SUM")
        self.assertEqual(self.history.cell(23, 12).value, 25)
        self.assertTrue(self.workbook.calculation.fullCalcOnLoad)


if __name__ == "__main__":
    unittest.main()
